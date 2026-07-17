"""
MBT POS — Shared Report Export Service
Professional XLSX / CSV formatting used by Reports, Consumption, Debt, Inventory.

Standards:
  • Header block: business name, optional logo, title, generated at/by, filters, period
  • Frozen header row + auto-filter on data tables
  • Currency: KSh / KES  #,##0.00
  • Dates: dd MMM yyyy (and dd MMM yyyy HH:mm for datetimes)
  • Footer: grand totals, record count, app version
  • CSV: UTF-8 with BOM for Excel
"""
from __future__ import annotations

import csv
import os
import sys
from datetime import datetime, date
from typing import Any, Iterable, Optional, Sequence, Union

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

# ── Brand palette ─────────────────────────────────────────────────────────────
D = "0D1B2A"
M = "1A3C5E"
G = "F4A825"
W = "FFFFFF"
ALT = "EEF3F8"
TOT_BG = "FFF3CC"
TOT_FG = "0D1B2A"
BDR = "B0BEC5"
GRN = "43A047"
RED = "E53935"
BLU = "1565C0"

CURRENCY_FMT = '#,##0.00'
DATE_FMT = 'dd MMM yyyy'
DATETIME_FMT = 'dd MMM yyyy HH:mm'
INT_FMT = '#,##0'
QTY_FMT = '#,##0.##'

HEADER_ROWS = 6  # rows 1–6 = brand / meta; data header typically at row 7


def _side(style='thin', color=BDR):
    return Side(style=style, color=color)


def _border(color=BDR):
    s = _side(color=color)
    return Border(left=s, right=s, top=s, bottom=s)


def _fill(hex_color: str):
    return PatternFill('solid', fgColor=hex_color)


def _font(size=10, bold=False, color='000000', italic=False, name='Calibri'):
    return Font(name=name, size=size, bold=bold, color=color, italic=italic)


def _align(h='left', v='center', wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


def app_version() -> str:
    try:
        root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        import json
        with open(os.path.join(root, 'version.json'), encoding='utf-8-sig') as f:
            return str(json.load(f).get('version') or '—')
    except Exception:
        return '—'


def currency_number_format(currency: str = 'KES') -> str:
    """Excel number format with currency code prefix (KSh preferred for Kenya)."""
    code = (currency or 'KES').strip()
    if code.upper() in ('KES', 'KSH', 'KSHS'):
        code = 'KSh'
    return f'"{code} "#,##0.00'


def get_export_dir() -> str:
    for d in (
        os.path.join(os.path.expanduser('~'), 'Desktop'),
        os.path.join(os.path.expanduser('~'), 'Documents'),
        os.path.expanduser('~'),
    ):
        if os.path.isdir(d):
            folder = os.path.join(d, 'MBT POS Exports')
            os.makedirs(folder, exist_ok=True)
            return folder
    try:
        from mbt_paths import get_project_root
        folder = os.path.join(get_project_root(), 'exports')
    except Exception:
        folder = os.path.join(
            os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
            'exports',
        )
    os.makedirs(folder, exist_ok=True)
    return folder


def find_logo_path() -> Optional[str]:
    """Locate brand logo for embedding in workbook header."""
    roots = []
    try:
        roots.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
    except Exception:
        pass
    if getattr(sys, 'frozen', False):
        roots.append(getattr(sys, '_MEIPASS', ''))
        roots.append(os.path.dirname(sys.executable))
    names = ('mbt_logo_hd.png', 'mbt_icon_256.png', 'mbt_icon.png', 'assets/mbt_logo_hd.png')
    for root in roots:
        if not root:
            continue
        for name in names:
            p = os.path.join(root, name)
            if os.path.isfile(p):
                return p
            p2 = os.path.join(root, 'assets', os.path.basename(name))
            if os.path.isfile(p2):
                return p2
            p3 = os.path.join(root, 'desktop', 'assets', os.path.basename(name))
            if os.path.isfile(p3):
                return p3
    return None


def format_date_value(val: Any) -> Any:
    """Return a datetime/date for Excel, or leave string/None as-is."""
    if val is None or val == '':
        return None
    if isinstance(val, datetime):
        return val
    if isinstance(val, date):
        return val
    s = str(val).strip().replace('T', ' ')
    if not s:
        return None
    # Strip fractional seconds / timezone for strptime
    core = s[:19] if len(s) >= 19 and s[4] == '-' else s
    for fmt, slice_len in (
        ('%Y-%m-%d %H:%M:%S', 19),
        ('%Y-%m-%d %H:%M', 16),
        ('%Y-%m-%d', 10),
    ):
        try:
            parsed = datetime.strptime(core[:slice_len], fmt)
            if fmt == '%Y-%m-%d':
                return parsed.date()
            return parsed
        except Exception:
            continue
    return s


def write_report_header(
    ws: Worksheet,
    *,
    shop_name: str,
    title: str,
    ncols: int,
    period: Optional[str] = None,
    generated_by: Optional[str] = None,
    filters: Optional[str] = None,
    currency: str = 'KES',
    logo_path: Optional[str] = None,
) -> int:
    """
    Write professional header block into rows 1–6.
    Returns the 1-based row index for the column header row (typically 7).
    """
    end = get_column_letter(max(1, ncols))
    now = datetime.now()

    # Row 1 — brand + title
    ws.merge_cells(f'A1:{end}1')
    c1 = ws['A1']
    c1.value = f'  {shop_name or "MBT POS"}  ·  {title}'
    c1.font = _font(16, bold=True, color=W)
    c1.fill = _fill(D)
    c1.alignment = _align('center', 'center')
    ws.row_dimensions[1].height = 32

    # Row 2 — vendor line + period
    ws.merge_cells(f'A2:{end}2')
    c2 = ws['A2']
    sub = 'MugoByte Technologies  ·  mugobyte.com'
    if period:
        sub += f'  ·  Period: {period}'
    c2.value = sub
    c2.font = _font(10, italic=True, color=G)
    c2.fill = _fill(M)
    c2.alignment = _align('center', 'center')
    ws.row_dimensions[2].height = 18

    # Row 3 — generated meta
    ws.merge_cells(f'A3:{end}3')
    meta_parts = [
        f"Generated: {now.strftime('%d %b %Y %H:%M')}",
        f"By: {generated_by or 'System'}",
        f"Version: {app_version()}",
        f"Currency: {currency}",
    ]
    ws['A3'].value = '  ·  '.join(meta_parts)
    ws['A3'].font = _font(9, color='37474F')
    ws['A3'].fill = _fill(ALT)
    ws['A3'].alignment = _align('left', 'center')
    ws.row_dimensions[3].height = 18

    # Row 4 — filters
    ws.merge_cells(f'A4:{end}4')
    ws['A4'].value = f"Filters: {filters}" if filters else 'Filters: (none)'
    ws['A4'].font = _font(9, italic=True, color='546E7A')
    ws['A4'].alignment = _align('left', 'center')
    ws.row_dimensions[4].height = 16

    # Row 5 — spacer
    ws.row_dimensions[5].height = 8

    # Optional logo (top-left overlay — keep small so it doesn't clash with title)
    path = logo_path if logo_path is not None else find_logo_path()
    if path and os.path.isfile(path):
        try:
            img = XLImage(path)
            # Keep logo modest
            img.width = 48
            img.height = 48
            ws.add_image(img, 'A1')
        except Exception:
            pass

    ws.sheet_view.showGridLines = False
    return 7  # column header row


def style_header_row(ws: Worksheet, row: int, headers: Sequence[str], widths: Optional[Sequence[float]] = None):
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.font = _font(11, bold=True, color=W)
        cell.fill = _fill(M)
        cell.alignment = _align('center', 'center')
        cell.border = _border()
        if widths and col <= len(widths):
            ws.column_dimensions[get_column_letter(col)].width = widths[col - 1]


def apply_data_cell(
    cell,
    value: Any,
    *,
    kind: str = 'text',
    currency: str = 'KES',
    alt: bool = False,
):
    """kind: text | currency | int | qty | date | datetime | pct | center"""
    if kind in ('date', 'datetime'):
        value = format_date_value(value)
    cell.value = value
    cell.border = _border()
    if alt:
        cell.fill = _fill(ALT)

    if kind == 'currency':
        cell.number_format = currency_number_format(currency)
        cell.alignment = _align('right')
    elif kind == 'int':
        cell.number_format = INT_FMT
        cell.alignment = _align('right')
    elif kind == 'qty':
        cell.number_format = QTY_FMT
        cell.alignment = _align('right')
    elif kind == 'pct':
        cell.number_format = '0.0%'
        cell.alignment = _align('right')
    elif kind == 'date':
        if isinstance(value, (datetime, date)):
            cell.number_format = DATE_FMT
        cell.alignment = _align('center')
    elif kind == 'datetime':
        if isinstance(value, datetime):
            cell.number_format = DATETIME_FMT
        elif isinstance(value, date):
            cell.number_format = DATE_FMT
        cell.alignment = _align('center')
    elif kind == 'center':
        cell.alignment = _align('center')
    else:
        cell.alignment = _align('left')


def write_totals_row(
    ws: Worksheet,
    row: int,
    ncols: int,
    *,
    label: str = 'GRAND TOTAL',
    values: Optional[dict] = None,
    currency: str = 'KES',
):
    """
    values: {col_index: (value, kind)} where kind is currency|int|qty|text
    """
    for col in range(1, ncols + 1):
        c = ws.cell(row=row, column=col)
        c.fill = _fill(TOT_BG)
        c.font = _font(10, bold=True, color=TOT_FG)
        c.border = _border()
    ws.cell(row=row, column=1).value = label
    if values:
        for col, (val, kind) in values.items():
            c = ws.cell(row=row, column=col, value=val)
            c.fill = _fill(TOT_BG)
            c.font = _font(10, bold=True, color=TOT_FG)
            c.border = _border()
            if kind == 'currency':
                c.number_format = currency_number_format(currency)
                c.alignment = _align('right')
            elif kind == 'int':
                c.number_format = INT_FMT
                c.alignment = _align('right')
            elif kind == 'qty':
                c.number_format = QTY_FMT
                c.alignment = _align('right')


def write_footer(
    ws: Worksheet,
    row: int,
    ncols: int,
    *,
    record_count: int = 0,
    extra: Optional[str] = None,
):
    end = get_column_letter(max(1, ncols))
    ws.merge_cells(f'A{row}:{end}{row}')
    parts = [
        f"Records: {record_count}",
        f"Generated: {datetime.now().strftime('%d %b %Y %H:%M')}",
        f"MBT POS v{app_version()}",
        'Powered by MugoByte Technologies · mugobyte.com',
    ]
    if extra:
        parts.insert(1, extra)
    c = ws[f'A{row}']
    c.value = '  |  '.join(parts)
    c.font = _font(9, italic=True, color=G)
    c.alignment = _align('center')


def finalize_table(
    ws: Worksheet,
    header_row: int,
    data_start: int,
    data_end: int,
    ncols: int,
):
    """Freeze panes below header + enable auto-filter on the data range."""
    ws.freeze_panes = f'A{header_row + 1}'
    if data_end >= data_start and ncols >= 1:
        end_col = get_column_letter(ncols)
        # AutoFilter includes header row through last data row
        ws.auto_filter.ref = f'A{header_row}:{end_col}{data_end}'


def new_workbook_sheet(title: str = 'Report') -> tuple:
    wb = Workbook()
    ws = wb.active
    ws.title = title[:31]
    return wb, ws


def save_workbook(wb: Workbook, output_path: str) -> str:
    os.makedirs(os.path.dirname(output_path) or '.', exist_ok=True)
    wb.save(output_path)
    return output_path


# ── Generic tabular export ────────────────────────────────────────────────────

def export_tabular_xlsx(
    *,
    title: str,
    headers: Sequence[str],
    rows: Sequence[Sequence[Any]],
    kinds: Optional[Sequence[str]] = None,
    widths: Optional[Sequence[float]] = None,
    shop_name: str = 'My Shop',
    period: Optional[str] = None,
    generated_by: Optional[str] = None,
    filters: Optional[str] = None,
    currency: str = 'KES',
    output_path: Optional[str] = None,
    filename: Optional[str] = None,
    total_cols: Optional[dict] = None,
    sheet_name: Optional[str] = None,
) -> str:
    """
    One-sheet professional export.
    kinds[i]: text|currency|int|qty|date|datetime|pct|center
    total_cols: {1-based col: (sum_value, kind)}
    """
    wb, ws = new_workbook_sheet(sheet_name or title)
    ncols = len(headers)
    header_row = write_report_header(
        ws,
        shop_name=shop_name,
        title=title,
        ncols=ncols,
        period=period,
        generated_by=generated_by,
        filters=filters,
        currency=currency,
    )
    style_header_row(ws, header_row, headers, widths)
    kinds = list(kinds or ['text'] * ncols)
    while len(kinds) < ncols:
        kinds.append('text')

    for ri, row in enumerate(rows):
        r = header_row + 1 + ri
        alt = ri % 2 == 1
        for ci in range(ncols):
            val = row[ci] if ci < len(row) else None
            apply_data_cell(
                ws.cell(row=r, column=ci + 1),
                val,
                kind=kinds[ci],
                currency=currency,
                alt=alt,
            )

    data_start = header_row + 1
    data_end = header_row + len(rows)
    if not rows:
        data_end = header_row  # no data rows

    totals_row = data_end + 1 if rows else header_row + 1
    if total_cols:
        write_totals_row(
            ws, totals_row, ncols,
            values=total_cols,
            currency=currency,
        )
        footer_row = totals_row + 2
    else:
        footer_row = (data_end if rows else header_row) + 2

    write_footer(ws, footer_row, ncols, record_count=len(rows))
    if rows:
        finalize_table(ws, header_row, data_start, data_end, ncols)
    else:
        ws.freeze_panes = f'A{header_row + 1}'

    if not output_path:
        fname = filename or f"MBT_Report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        output_path = os.path.join(get_export_dir(), fname)
    return save_workbook(wb, output_path)


def export_csv(
    *,
    headers: Sequence[str],
    rows: Iterable[Sequence[Any]],
    output_path: str,
) -> str:
    """UTF-8 BOM CSV so Excel opens accented characters correctly."""
    os.makedirs(os.path.dirname(output_path) or '.', exist_ok=True)
    with open(output_path, 'w', encoding='utf-8-sig', newline='') as f:
        w = csv.writer(f)
        w.writerow(list(headers))
        for row in rows:
            out = []
            for v in row:
                if isinstance(v, datetime):
                    out.append(v.strftime('%d %b %Y %H:%M'))
                elif isinstance(v, date):
                    out.append(v.strftime('%d %b %Y'))
                elif v is None:
                    out.append('')
                else:
                    out.append(v)
            w.writerow(out)
    return output_path


# ── Domain helpers used by tabs ───────────────────────────────────────────────

def export_consumption_report(
    rows: Sequence[dict],
    *,
    shop_name: str = 'My Shop',
    start_date: str,
    end_date: str,
    currency: str = 'KES',
    generated_by: Optional[str] = None,
    filters: Optional[str] = None,
    totals: Optional[dict] = None,
    output_path: Optional[str] = None,
) -> str:
    headers = [
        'Date', 'Reference', 'Department', 'Taken By', 'Reason', 'Notes',
        'Product', 'Qty', 'Unit Cost', 'Total Cost', 'User', 'Status',
    ]
    table = []
    for r in rows:
        voided = int(r.get('voided') or 0) == 1
        table.append([
            r.get('date') or '',
            r.get('reference_no') or '',
            r.get('department_name') or '',
            r.get('taken_by') or '',
            r.get('reason') or '',
            r.get('notes') or '',
            r.get('product_name') or '',
            float(r.get('quantity') or 0),
            float(r.get('unit_cost') or 0),
            float(r.get('total_cost') or 0),
            r.get('created_by_name') or '',
            'Voided' if voided else 'OK',
        ])
    tot = totals or {}
    total_cols = {
        8: (float(tot.get('total_qty') or sum(r[7] for r in table)), 'qty'),
        10: (float(tot.get('total_cost') or sum(r[9] for r in table)), 'currency'),
    }
    path = output_path or os.path.join(
        get_export_dir(),
        f'Internal_Consumption_{start_date}_to_{end_date}.xlsx',
    )
    return export_tabular_xlsx(
        title='Internal Consumption Report',
        headers=headers,
        rows=table,
        kinds=['date', 'text', 'text', 'text', 'text', 'text',
               'text', 'qty', 'currency', 'currency', 'text', 'center'],
        widths=[12, 14, 14, 14, 18, 24, 22, 10, 12, 14, 14, 10],
        shop_name=shop_name,
        period=f'{start_date} → {end_date}',
        generated_by=generated_by,
        filters=filters or f'Date range {start_date} to {end_date}',
        currency=currency,
        output_path=path,
        total_cols=total_cols,
        sheet_name='Internal Consumption',
    )


def export_debt_report(
    *,
    invoices: Sequence[dict],
    payments: Sequence[dict],
    aging: Optional[dict] = None,
    summary: Optional[dict] = None,
    shop_name: str = 'My Shop',
    currency: str = 'KES',
    generated_by: Optional[str] = None,
    filters: Optional[str] = None,
    period: Optional[str] = None,
    output_path: Optional[str] = None,
) -> str:
    wb = Workbook()

    # Sheet 1 — Invoices
    ws = wb.active
    ws.title = 'Debt Invoices'
    headers = [
        'Invoice No.', 'Customer', 'Phone', 'Total', 'Paid', 'Balance',
        'Status', 'Due Date', 'Created',
    ]
    hdr = write_report_header(
        ws, shop_name=shop_name, title='Debt Invoices',
        ncols=len(headers), period=period, generated_by=generated_by,
        filters=filters, currency=currency,
    )
    # KPI strip
    summary = summary or {}
    outstanding = float((summary.get('outstanding') or {}).get('total', 0) or 0)
    overdue = float((summary.get('overdue') or {}).get('total', 0) or 0)
    ws.cell(5, 1).value = f"Outstanding: {outstanding:,.2f}   ·   Overdue: {overdue:,.2f}   ·   Open invoices: {len([i for i in invoices if i.get('status') not in ('paid','cancelled')])}"
    ws.cell(5, 1).font = _font(10, bold=True, color=M)

    style_header_row(ws, hdr, headers, [16, 22, 14, 14, 14, 14, 12, 14, 16])
    kinds = ['text', 'text', 'text', 'currency', 'currency', 'currency', 'center', 'date', 'datetime']
    open_inv = [i for i in invoices if i.get('status') not in ('paid', 'cancelled')]
    for ri, inv in enumerate(open_inv):
        r = hdr + 1 + ri
        vals = [
            inv.get('invoice_number', ''),
            inv.get('customer_name', ''),
            inv.get('customer_phone', '') or '',
            float(inv.get('total_amount') or 0),
            float(inv.get('amount_paid') or 0),
            float(inv.get('balance') or 0),
            (inv.get('status') or '').upper(),
            inv.get('due_date') or '',
            inv.get('created_at') or '',
        ]
        for ci, (v, k) in enumerate(zip(vals, kinds), 1):
            apply_data_cell(ws.cell(r, ci), v, kind=k, currency=currency, alt=ri % 2 == 1)
    data_end = hdr + len(open_inv)
    bal_sum = sum(float(i.get('balance') or 0) for i in open_inv)
    write_totals_row(
        ws, data_end + 1 if open_inv else hdr + 1, len(headers),
        values={6: (bal_sum, 'currency')},
        currency=currency,
    )
    write_footer(ws, (data_end + 3) if open_inv else hdr + 3, len(headers), record_count=len(open_inv))
    if open_inv:
        finalize_table(ws, hdr, hdr + 1, data_end, len(headers))

    # Sheet 2 — Aging
    ws2 = wb.create_sheet('Aging')
    aging = aging or {}
    ah = ['Bucket', 'Invoices', f'Amount ({currency})']
    h2 = write_report_header(
        ws2, shop_name=shop_name, title='Debt Aging',
        ncols=3, period=period, generated_by=generated_by,
        filters=filters, currency=currency,
    )
    style_header_row(ws2, h2, ah, [18, 12, 18])
    bands = [
        ('Current', aging.get('current', {})),
        ('1-30 Days', aging.get('1_30', {})),
        ('31-60 Days', aging.get('31_60', {})),
        ('61-90 Days', aging.get('61_90', {})),
        ('90+ Days', aging.get('over_90', {})),
    ]
    for i, (label, data) in enumerate(bands):
        r = h2 + 1 + i
        apply_data_cell(ws2.cell(r, 1), label, kind='text', alt=i % 2 == 1)
        apply_data_cell(ws2.cell(r, 2), int((data or {}).get('count') or 0), kind='int', alt=i % 2 == 1)
        apply_data_cell(ws2.cell(r, 3), float((data or {}).get('total') or 0), kind='currency', currency=currency, alt=i % 2 == 1)
    write_footer(ws2, h2 + 8, 3, record_count=5)
    finalize_table(ws2, h2, h2 + 1, h2 + 5, 3)

    # Sheet 3 — Payments
    ws3 = wb.create_sheet('Payments')
    ph = ['Receipt', 'Customer', 'Amount', 'Method', 'Date']
    h3 = write_report_header(
        ws3, shop_name=shop_name, title='Debt Payments',
        ncols=5, period=period, generated_by=generated_by,
        filters=filters, currency=currency,
    )
    style_header_row(ws3, h3, ph, [18, 24, 16, 14, 18])
    pay_rows = sorted(payments or [], key=lambda x: x.get('created_at') or '', reverse=True)
    for i, pay in enumerate(pay_rows):
        r = h3 + 1 + i
        vals = [
            pay.get('payment_receipt', ''),
            pay.get('customer_name', ''),
            float(pay.get('amount') or 0),
            (pay.get('payment_method') or '').upper(),
            pay.get('created_at') or '',
        ]
        kinds3 = ['text', 'text', 'currency', 'center', 'datetime']
        for ci, (v, k) in enumerate(zip(vals, kinds3), 1):
            apply_data_cell(ws3.cell(r, ci), v, kind=k, currency=currency, alt=i % 2 == 1)
    pe = h3 + len(pay_rows)
    pay_sum = sum(float(p.get('amount') or 0) for p in pay_rows)
    write_totals_row(
        ws3, pe + 1 if pay_rows else h3 + 1, 5,
        values={3: (pay_sum, 'currency')},
        currency=currency,
    )
    write_footer(ws3, (pe + 3) if pay_rows else h3 + 3, 5, record_count=len(pay_rows))
    if pay_rows:
        finalize_table(ws3, h3, h3 + 1, pe, 5)

    path = output_path or os.path.join(
        get_export_dir(),
        f"MBT_Debt_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
    )
    return save_workbook(wb, path)


def export_inventory_movements(
    movements: Sequence[dict],
    *,
    products_by_id: Optional[dict] = None,
    shop_name: str = 'My Shop',
    currency: str = 'KES',
    generated_by: Optional[str] = None,
    filters: Optional[str] = None,
    output_path: Optional[str] = None,
) -> str:
    products_by_id = products_by_id or {}
    headers = [
        'Date', 'Product', 'SKU', 'Type', 'Qty Change', 'Stock After',
        'Reason', 'User', 'Reference',
    ]
    table = []
    for m in movements:
        pid = m.get('product_id')
        prod = products_by_id.get(pid) or {}
        table.append([
            m.get('created_at') or '',
            prod.get('name') or m.get('product_name') or '',
            prod.get('sku') or m.get('sku') or '',
            m.get('movement_type') or '',
            float(m.get('quantity') or m.get('qty_change') or 0),
            float(m.get('stock_after') or m.get('quantity_after') or 0),
            m.get('reason') or m.get('notes') or '',
            m.get('user_name') or m.get('created_by_name') or '',
            m.get('reference') or m.get('reference_no') or '',
        ])
    path = output_path or os.path.join(
        get_export_dir(),
        f"MBT_Stock_Movements_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
    )
    return export_tabular_xlsx(
        title='Stock Movements',
        headers=headers,
        rows=table,
        kinds=['datetime', 'text', 'text', 'center', 'qty', 'qty', 'text', 'text', 'text'],
        widths=[18, 28, 14, 14, 12, 12, 22, 14, 16],
        shop_name=shop_name,
        period=f"As at {datetime.now().strftime('%d %b %Y %H:%M')}",
        generated_by=generated_by,
        filters=filters or 'All recent stock movements',
        currency=currency,
        output_path=path,
        sheet_name='Stock Movements',
    )


def export_inventory_full(
    products: Sequence[dict],
    movements: Sequence[dict],
    *,
    shop_name: str = 'My Shop',
    currency: str = 'KES',
    generated_by: Optional[str] = None,
    output_path: Optional[str] = None,
) -> str:
    """Snapshot + movements in one workbook."""
    products_by_id = {p.get('id'): p for p in products if p.get('id') is not None}
    snap = export_inventory_snapshot(
        products, shop_name=shop_name, currency=currency,
        generated_by=generated_by,
        output_path=None,  # temp path unused — we rebuild
    )
    # Rebuild as combined to keep formatting consistent
    from openpyxl import load_workbook
    import tempfile

    path = output_path or os.path.join(
        get_export_dir(),
        f"MBT_Inventory_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
    )
    # Start from snapshot file, append movements sheet content by re-exporting movements to temp
    move_tmp = os.path.join(
        get_export_dir(),
        f"_tmp_moves_{datetime.now().strftime('%H%M%S%f')}.xlsx",
    )
    export_inventory_movements(
        movements, products_by_id=products_by_id, shop_name=shop_name,
        currency=currency, generated_by=generated_by,
        filters=f'{len(movements)} recent movements',
        output_path=move_tmp,
    )
    wb = load_workbook(snap)
    wb2 = load_workbook(move_tmp)
    # Rename first sheet
    wb.active.title = 'Inventory'
    src = wb2.active
    dst = wb.create_sheet('Stock Movements')
    for row in src.iter_rows():
        for cell in row:
            nc = dst.cell(cell.row, cell.column, cell.value)
            if cell.has_style:
                nc.font = cell.font.copy()
                nc.fill = cell.fill.copy()
                nc.border = cell.border.copy()
                nc.alignment = cell.alignment.copy()
                nc.number_format = cell.number_format
    for col, dim in src.column_dimensions.items():
        dst.column_dimensions[col].width = dim.width
    if src.freeze_panes:
        dst.freeze_panes = src.freeze_panes
    if src.auto_filter and src.auto_filter.ref:
        dst.auto_filter.ref = src.auto_filter.ref
    for merged in src.merged_cells.ranges:
        try:
            dst.merge_cells(str(merged))
        except Exception:
            pass
    wb.save(path)
    for p in (snap, move_tmp):
        try:
            os.remove(p)
        except Exception:
            pass
    return path


def export_inventory_snapshot(
    products: Sequence[dict],
    *,
    shop_name: str = 'My Shop',
    currency: str = 'KES',
    generated_by: Optional[str] = None,
    output_path: Optional[str] = None,
) -> str:
    headers = [
        '#', 'Product Name', 'SKU', 'Category', 'Selling Price', 'Cost Price',
        'Current Stock', 'Min Stock', 'Stock Value', 'Status',
    ]
    rows = []
    total_value = 0.0
    for idx, prod in enumerate(products, 1):
        price = float(prod.get('price') or 0)
        cost = float(prod.get('cost_price') or 0)
        stock = float(prod.get('stock') or 0)
        min_stock = float(prod.get('min_stock') or 5)
        stk_value = cost * stock if cost else price * stock * 0.6
        total_value += stk_value
        if stock <= 0:
            status = 'OUT OF STOCK'
        elif stock <= min_stock:
            status = 'LOW STOCK'
        else:
            status = 'OK'
        rows.append([
            idx, prod.get('name', ''), prod.get('sku') or '',
            prod.get('category') or '', price, cost, stock, min_stock,
            stk_value, status,
        ])
    path = output_path or os.path.join(
        get_export_dir(),
        f"MBT_Inventory_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
    )
    return export_tabular_xlsx(
        title='Inventory Snapshot',
        headers=headers,
        rows=rows,
        kinds=['int', 'text', 'text', 'text', 'currency', 'currency',
               'qty', 'qty', 'currency', 'center'],
        widths=[5, 30, 14, 16, 14, 14, 12, 12, 14, 14],
        shop_name=shop_name,
        period=f"As at {datetime.now().strftime('%d %b %Y %H:%M')}",
        generated_by=generated_by,
        filters='Current stock levels',
        currency=currency,
        output_path=path,
        total_cols={9: (total_value, 'currency')},
        sheet_name='Inventory',
    )
