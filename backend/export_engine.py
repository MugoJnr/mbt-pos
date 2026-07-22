"""
MBT POS - Spreadsheet Export Engine
MugoByte Technologies | mugobyte.com

Multi-sheet Excel sales report built on report_export_service
(professional headers, freeze panes, auto-filter, KSh formats, version footer).
"""
import os
import sys
from datetime import datetime, date
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.utils import get_column_letter

from backend.report_export_service import (
    ALT, BDR, BLU, D, G, GRN, M, RED, TOT_BG, TOT_FG, W,
    _align, _border, _fill, _font,
    apply_data_cell, currency_number_format, find_logo_path,
    finalize_table, get_export_dir, write_footer, write_report_header,
    write_totals_row, style_header_row, app_version,
)


def _hdr_cell(ws, row, col, value, width=None, col_letter=None):
    c = ws.cell(row=row, column=col, value=value)
    c.font = _font(11, bold=True, color=W)
    c.fill = _fill(M)
    c.alignment = _align('center')
    c.border = _border(BDR)
    if width and col_letter:
        ws.column_dimensions[col_letter].width = width
    return c


def _kpi_block(ws, kpis, start_row=4):
    ws.cell(start_row - 1, 1).value = "KEY PERFORMANCE INDICATORS"
    ws.cell(start_row - 1, 1).font = _font(12, bold=True, color=M)
    ws.row_dimensions[start_row - 1].height = 22
    for i, (lbl, val, fmt) in enumerate(kpis):
        r = start_row + i
        lc = ws.cell(r, 1, lbl)
        lc.font = _font(10, bold=True)
        lc.fill = _fill(ALT)
        lc.border = _border()
        lc.alignment = _align()
        vc = ws.cell(r, 2, val)
        vc.border = _border()
        vc.alignment = _align('right')
        vc.font = _font(10, bold=True, color=BLU)
        if fmt == 'currency':
            vc.number_format = '#,##0.00'
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 20


def export_sales_report(
    sales_data, items_by_sale, shop_name='My Shop',
    start_date=None, end_date=None, output_path=None,
    currency='KES', products_data=None,
    debt_summary=None, aging_report=None,
    debt_invoices=None, debt_payments=None,
    variance_rows=None, variance_summary=None,
    generated_by=None, filters=None,
):
    """
    Multi-sheet Excel report.
    Sheet 1 – Sales Summary
    Sheet 2 – Line Items
    Sheet 3 – Top Products (+ chart)
    Sheet 4 – Payment Methods
    Sheet 5 – Stock / Inventory
    Sheet 6 – Debt Management
    Sheet 7 – Payment Variance (optional)
    """
    wb = Workbook()
    period = f"{start_date or '—'} to {end_date or str(date.today())}"
    filt = filters or f"Completed sales · {period}"
    who = generated_by or 'System'
    cur_fmt = currency_number_format(currency)

    total_rev = sum(float(s.get('total', 0) or 0) for s in sales_data)
    total_count = len(sales_data)
    avg_sale = total_rev / total_count if total_count else 0
    total_disc = sum(float(s.get('discount', 0) or 0) for s in sales_data)
    total_tax = sum(float(s.get('tax', 0) or 0) for s in sales_data)
    total_round = sum(float(s.get('cash_rounding_adj', 0) or 0) for s in sales_data)
    total_orig = 0.0
    for s in sales_data:
        tot = float(s.get('total', 0) or 0)
        adj = float(s.get('cash_rounding_adj', 0) or 0)
        orig = float(s.get('original_total') or 0)
        if orig <= 0:
            orig = tot - adj
        total_orig += orig

    line_items = []
    for sale in sales_data:
        sid = sale.get('id') or sale.get('sale_id')
        items = items_by_sale.get(sid, [])
        if not items and 'items' in sale:
            items = sale['items']
        for item in items:
            line_items.append({
                'receipt': sale.get('receipt_number', ''),
                'date': sale.get('created_at', '') or '',
                'cashier': sale.get('cashier_name', ''),
                'product_name': item.get('product_name', ''),
                'sku': item.get('sku', '') or '',
                'quantity': float(item.get('quantity', 0) or 0),
                'unit_price': float(item.get('unit_price', 0) or 0),
                'discount': float(item.get('discount', 0) or 0),
                'total': float(item.get('total', 0) or 0),
                'payment': (sale.get('payment_method', '') or '').upper(),
            })

    # ── Sheet 1 – Sales Summary ───────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Sales Summary"
    hdr1 = write_report_header(
        ws1, shop_name=shop_name, title="Sales Report", ncols=12,
        period=period, generated_by=who, filters=filt, currency=currency,
    )
    # KPIs sit in rows 5–10 area; move table header after KPIs
    _kpi_block(ws1, [
        ('Total Transactions', total_count, ''),
        (f'Total Revenue ({currency})', total_rev, 'currency'),
        (f'Original Total ({currency})', total_orig, 'currency'),
        (f'Cash Rounding ({currency})', total_round, 'currency'),
        (f'Average Sale ({currency})', avg_sale, 'currency'),
        (f'Total Discounts ({currency})', total_disc, 'currency'),
        (f'Total Tax ({currency})', total_tax, 'currency'),
    ], start_row=5)

    SR = 14
    ws1.cell(SR - 1, 1).value = "TRANSACTION DETAILS"
    ws1.cell(SR - 1, 1).font = _font(12, bold=True, color=M)

    hdrs1 = [
        '#', 'Receipt No.', 'Date & Time', 'Cashier', 'Items',
        f'Subtotal ({currency})', f'Discount ({currency})',
        f'Tax ({currency})', f'Original ({currency})',
        f'Rounding ({currency})', f'Final ({currency})', 'Payment',
    ]
    wds1 = [5, 18, 20, 16, 6, 16, 14, 12, 14, 12, 16, 12]
    style_header_row(ws1, SR, hdrs1, wds1)

    for idx, sale in enumerate(sales_data):
        r = SR + 1 + idx
        sid = sale.get('id') or sale.get('sale_id')
        n_items = len(items_by_sale.get(sid, sale.get('items', [])))
        adj = float(sale.get('cash_rounding_adj') or 0)
        tot = float(sale.get('total', 0) or 0)
        orig = float(sale.get('original_total') or 0)
        if orig <= 0:
            orig = tot - adj
        vals = [
            idx + 1,
            sale.get('receipt_number', ''),
            sale.get('created_at', '') or '',
            sale.get('cashier_name', ''),
            n_items or sale.get('item_count', '') or 0,
            float(sale.get('subtotal', sale.get('total', 0)) or 0),
            float(sale.get('discount', 0) or 0),
            float(sale.get('tax', 0) or 0),
            orig, adj, tot,
            (sale.get('payment_method', '') or '').upper(),
        ]
        kinds = [
            'int', 'text', 'datetime', 'text', 'int',
            'currency', 'currency', 'currency', 'currency', 'currency', 'currency', 'center',
        ]
        for col, (val, kind) in enumerate(zip(vals, kinds), 1):
            apply_data_cell(
                ws1.cell(r, col), val, kind=kind, currency=currency, alt=idx % 2 == 1)

    TR1 = SR + 1 + len(sales_data)
    write_totals_row(ws1, TR1, 12, currency=currency, values={
        6: (sum(float(s.get('subtotal', s.get('total', 0)) or 0) for s in sales_data), 'currency'),
        7: (total_disc, 'currency'),
        8: (total_tax, 'currency'),
        9: (total_orig, 'currency'),
        10: (total_round, 'currency'),
        11: (total_rev, 'currency'),
    })
    write_footer(ws1, TR1 + 2, 12, record_count=total_count,
                 extra=f"Revenue {total_rev:,.2f} · Rounding {total_round:+,.2f}")
    if sales_data:
        finalize_table(ws1, SR, SR + 1, SR + total_count, 12)

    # ── Sheet 2 – Line Items ──────────────────────────────────────────────────
    ws2 = wb.create_sheet("Line Items")
    hdr2 = write_report_header(
        ws2, shop_name=shop_name, title="Sales Line Items — Full Detail",
        ncols=10, period=period, generated_by=who, filters=filt, currency=currency,
    )
    hdrs2 = [
        'Receipt No.', 'Date & Time', 'Cashier', 'Product Name', 'SKU',
        'Qty', f'Unit Price ({currency})', f'Discount ({currency})',
        f'Total ({currency})', 'Payment',
    ]
    style_header_row(ws2, hdr2, hdrs2, [18, 20, 16, 30, 14, 8, 16, 14, 16, 12])
    for idx, li in enumerate(line_items):
        r = hdr2 + 1 + idx
        vals = [
            li['receipt'], li['date'], li['cashier'], li['product_name'], li['sku'],
            li['quantity'], li['unit_price'], li['discount'], li['total'], li['payment'],
        ]
        kinds = [
            'text', 'datetime', 'text', 'text', 'text',
            'qty', 'currency', 'currency', 'currency', 'center',
        ]
        for col, (val, kind) in enumerate(zip(vals, kinds), 1):
            apply_data_cell(
                ws2.cell(r, col), val, kind=kind, currency=currency, alt=idx % 2 == 1)
    TR2 = hdr2 + 1 + len(line_items)
    write_totals_row(ws2, TR2, 10, currency=currency, values={
        6: (sum(li['quantity'] for li in line_items), 'qty'),
        8: (sum(li['discount'] for li in line_items), 'currency'),
        9: (sum(li['total'] for li in line_items), 'currency'),
    })
    write_footer(ws2, TR2 + 2, 10, record_count=len(line_items))
    if line_items:
        finalize_table(ws2, hdr2, hdr2 + 1, hdr2 + len(line_items), 10)

    # ── Sheet 3 – Top Products ────────────────────────────────────────────────
    ws3 = wb.create_sheet("Top Products")
    hdr3 = write_report_header(
        ws3, shop_name=shop_name, title="Product Performance",
        ncols=7, period=period, generated_by=who, filters=filt, currency=currency,
    )
    prod_stats = {}
    for li in line_items:
        name = li['product_name'] or 'Unknown'
        prod_stats.setdefault(name, {'qty': 0, 'revenue': 0, 'transactions': 0})
        prod_stats[name]['qty'] += li['quantity']
        prod_stats[name]['revenue'] += li['total']
        prod_stats[name]['transactions'] += 1
    sorted_prods = sorted(prod_stats.items(), key=lambda x: x[1]['revenue'], reverse=True)
    hdrs3 = [
        'Rank', 'Product Name', 'Units Sold', 'Transactions',
        f'Revenue ({currency})', '% of Total', 'Avg Unit Price',
    ]
    style_header_row(ws3, hdr3, hdrs3, [6, 32, 14, 14, 18, 14, 18])
    grand_rev = sum(v['revenue'] for v in prod_stats.values()) or 1
    for rank, (name, st) in enumerate(sorted_prods, 1):
        r = hdr3 + rank
        avg_up = st['revenue'] / st['qty'] if st['qty'] else 0
        vals = [
            rank, name, st['qty'], st['transactions'],
            st['revenue'], st['revenue'] / grand_rev, avg_up,
        ]
        kinds = ['int', 'text', 'qty', 'int', 'currency', 'pct', 'currency']
        for col, (val, kind) in enumerate(zip(vals, kinds), 1):
            apply_data_cell(
                ws3.cell(r, col), val, kind=kind, currency=currency, alt=rank % 2 == 0)
    TR3 = hdr3 + len(sorted_prods) + 1
    write_totals_row(ws3, TR3, 7, currency=currency, values={
        5: (grand_rev if prod_stats else 0, 'currency'),
    })
    if sorted_prods:
        chart = BarChart()
        chart.type = "bar"
        chart.title = "Revenue by Product"
        chart.y_axis.title = f"Revenue ({currency})"
        top_n = min(10, len(sorted_prods))
        data_ref = Reference(ws3, min_col=5, min_row=hdr3, max_row=hdr3 + top_n)
        cats_ref = Reference(ws3, min_col=2, min_row=hdr3 + 1, max_row=hdr3 + top_n)
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)
        chart.height = 14
        chart.width = 24
        ws3.add_chart(chart, "I5")
        finalize_table(ws3, hdr3, hdr3 + 1, hdr3 + len(sorted_prods), 7)
    write_footer(ws3, TR3 + 2, 7, record_count=len(sorted_prods))

    # ── Sheet 4 – Payment Methods ─────────────────────────────────────────────
    ws4 = wb.create_sheet("Payment Methods")
    hdr4 = write_report_header(
        ws4, shop_name=shop_name, title="Payment Method Breakdown",
        ncols=5, period=period, generated_by=who, filters=filt, currency=currency,
    )
    pay_stats = {}
    for sale in sales_data:
        pm = (sale.get('payment_method') or 'cash').upper()
        pay_stats.setdefault(pm, {'count': 0, 'total': 0.0})
        pay_stats[pm]['count'] += 1
        pay_stats[pm]['total'] += float(sale.get('total', 0) or 0)
    hdrs4 = [
        'Payment Method', 'Transactions', '% Transactions',
        f'Revenue ({currency})', '% Revenue',
    ]
    style_header_row(ws4, hdr4, hdrs4, [22, 16, 18, 22, 16])
    grand_count = sum(v['count'] for v in pay_stats.values()) or 1
    grand_pay = sum(v['total'] for v in pay_stats.values()) or 1
    for idx, (pm, st) in enumerate(
            sorted(pay_stats.items(), key=lambda x: x[1]['total'], reverse=True)):
        r = hdr4 + 1 + idx
        vals = [
            pm, st['count'], st['count'] / grand_count,
            st['total'], st['total'] / grand_pay,
        ]
        kinds = ['text', 'int', 'pct', 'currency', 'pct']
        for col, (val, kind) in enumerate(zip(vals, kinds), 1):
            apply_data_cell(
                ws4.cell(r, col), val, kind=kind, currency=currency, alt=idx % 2 == 1)
    write_footer(ws4, hdr4 + 2 + len(pay_stats), 5, record_count=len(pay_stats))
    if pay_stats:
        finalize_table(ws4, hdr4, hdr4 + 1, hdr4 + len(pay_stats), 5)

    # ── Sheet 5 – Stock / Inventory ───────────────────────────────────────────
    ws5 = wb.create_sheet("Stock & Inventory")
    hdr5 = write_report_header(
        ws5, shop_name=shop_name, title="Stock & Inventory Snapshot",
        ncols=10,
        period=f"As at {datetime.now().strftime('%d %b %Y %H:%M')}",
        generated_by=who, filters="Current stock levels", currency=currency,
    )
    ws5.cell(5, 1).value = (
        "Current stock at export — RED rows = below minimum (reorder needed)"
    )
    ws5.cell(5, 1).font = _font(10, italic=True, color=G)

    hdrs5 = [
        '#', 'Product Name', 'SKU', 'Category',
        f'Selling Price ({currency})', f'Cost Price ({currency})',
        'Current Stock', 'Min Stock', f'Stock Value ({currency})', 'Status',
    ]
    style_header_row(ws5, hdr5, hdrs5, [5, 30, 14, 16, 16, 14, 12, 10, 16, 14])

    LOW_FILL = _fill('FFEBEE')
    LOW_FONT = _font(10, bold=True, color=RED)
    OK_FONT = _font(10, bold=True, color=GRN)
    ZERO_FILL = _fill('FFF3E0')
    ZERO_FONT = _font(10, bold=True, color='E65100')

    products = products_data or []

    def _sort_key(p):
        stk = float(p.get('stock', 0) or 0)
        mn = float(p.get('min_stock', 5) or 5)
        if stk == 0:
            return (0, p.get('name', ''))
        if stk <= mn:
            return (1, p.get('name', ''))
        return (2, p.get('name', ''))

    products_sorted = sorted(products, key=_sort_key)
    total_stock_value = 0.0
    low_count = 0
    zero_count = 0

    for idx, prod in enumerate(products_sorted):
        r = hdr5 + 1 + idx
        alt = idx % 2 == 1
        name = prod.get('name', '')
        sku = prod.get('sku', '') or ''
        category = prod.get('category', '') or ''
        price = float(prod.get('price', 0) or 0)
        cost = float(prod.get('cost_price', 0) or 0)
        stock = float(prod.get('stock', 0) or 0)
        min_stock = float(prod.get('min_stock', 5) or 5)
        stk_value = cost * stock if cost else price * stock * 0.6
        total_stock_value += stk_value

        if stock == 0:
            status = 'OUT OF STOCK'
            zero_count += 1
        elif stock <= min_stock:
            status = 'LOW STOCK'
            low_count += 1
        else:
            status = 'OK'

        vals = [idx + 1, name, sku, category, price, cost, stock, min_stock, stk_value, status]
        kinds = ['int', 'text', 'text', 'text', 'currency', 'currency',
                 'qty', 'qty', 'currency', 'center']
        for col, (val, kind) in enumerate(zip(vals, kinds), 1):
            cell = ws5.cell(r, col)
            apply_data_cell(cell, val, kind=kind, currency=currency, alt=alt)
            if stock == 0:
                cell.fill = ZERO_FILL
            elif stock <= min_stock:
                cell.fill = LOW_FILL
            if col == 10:
                cell.font = (
                    ZERO_FONT if stock == 0 else
                    LOW_FONT if stock <= min_stock else OK_FONT
                )

    n_prods = len(products_sorted)
    TR5 = hdr5 + 1 + n_prods
    write_totals_row(ws5, TR5, 10, label='TOTALS', currency=currency, values={
        9: (total_stock_value, 'currency'),
    })
    ws5.cell(TR5, 2).value = f'{n_prods} products · {zero_count} out · {low_count} low'
    ws5.cell(TR5, 2).font = _font(10, bold=True, color=TOT_FG)
    ws5.cell(TR5, 2).fill = _fill(TOT_BG)

    write_footer(
        ws5, TR5 + 2, 10, record_count=n_prods,
        extra=f"Stock value {total_stock_value:,.2f}",
    )
    if products_sorted:
        finalize_table(ws5, hdr5, hdr5 + 1, hdr5 + n_prods, 10)

    # ── Sheet 6 – Debt Management ─────────────────────────────────────────────
    ws6 = wb.create_sheet("Debt Management")
    debt_summary = debt_summary or {}
    aging_report = aging_report or {}
    debt_invoices = debt_invoices or []
    debt_payments = debt_payments or []

    hdr6 = write_report_header(
        ws6, shop_name=shop_name, title="Debt Management Report",
        ncols=9, period=period, generated_by=who, filters=filt, currency=currency,
    )
    outstanding = float((debt_summary.get('outstanding') or {}).get('total', 0) or 0)
    overdue = float((debt_summary.get('overdue') or {}).get('total', 0) or 0)
    collected = float((debt_summary.get('today_collected') or {}).get('total', 0) or 0)
    customers_with_debt = int(debt_summary.get('customers_with_debt') or 0)

    _kpi_block(ws6, [
        (f'Outstanding Debt ({currency})', outstanding, 'currency'),
        (f'Overdue Debt ({currency})', overdue, 'currency'),
        (f'Collected Today ({currency})', collected, 'currency'),
        ('Customers with Debt', customers_with_debt, ''),
        ('Open Invoices', len([
            i for i in debt_invoices if i.get('status') not in ('paid', 'cancelled')
        ]), ''),
    ], start_row=5)

    ws6.cell(5, 4).value = "AGING BREAKDOWN"
    ws6.cell(5, 4).font = _font(12, bold=True, color=M)
    aging_rows = [
        ('Current', aging_report.get('current', {})),
        ('1-30 Days', aging_report.get('1_30', {})),
        ('31-60 Days', aging_report.get('31_60', {})),
        ('61-90 Days', aging_report.get('61_90', {})),
        ('90+ Days', aging_report.get('over_90', {})),
    ]
    _hdr_cell(ws6, 6, 4, "Bucket", 14, 'D')
    _hdr_cell(ws6, 6, 5, "Invoices", 10, 'E')
    _hdr_cell(ws6, 6, 6, f"Amount ({currency})", 16, 'F')
    for i, (label, data) in enumerate(aging_rows):
        r = 7 + i
        apply_data_cell(ws6.cell(r, 4), label, kind='text', alt=i % 2 == 1)
        apply_data_cell(
            ws6.cell(r, 5), int((data or {}).get('count', 0) or 0),
            kind='int', alt=i % 2 == 1)
        apply_data_cell(
            ws6.cell(r, 6), float((data or {}).get('total', 0) or 0),
            kind='currency', currency=currency, alt=i % 2 == 1)

    INV_HDR = 14
    ws6.cell(INV_HDR - 1, 1).value = "OPEN / RECENT DEBT INVOICES"
    ws6.cell(INV_HDR - 1, 1).font = _font(12, bold=True, color=M)
    hdrs6 = [
        'Invoice No.', 'Customer', f'Outstanding ({currency})',
        'Status', 'Due Date', 'Created',
    ]
    style_header_row(ws6, INV_HDR, hdrs6, [18, 24, 18, 12, 14, 20])
    inv_rows = sorted(
        debt_invoices, key=lambda x: x.get('created_at', ''), reverse=True
    )[:200]
    for i, inv in enumerate(inv_rows):
        r = INV_HDR + 1 + i
        vals = [
            inv.get('invoice_number', ''),
            inv.get('customer_name', ''),
            float(inv.get('balance', 0) or 0),
            (inv.get('status', '') or '').upper(),
            inv.get('due_date') or '',
            inv.get('created_at', '') or '',
        ]
        kinds = ['text', 'text', 'currency', 'center', 'date', 'datetime']
        for col, (val, kind) in enumerate(zip(vals, kinds), 1):
            apply_data_cell(
                ws6.cell(r, col), val, kind=kind, currency=currency, alt=i % 2 == 1)

    pay_start = INV_HDR + 2 + len(inv_rows)
    ws6.cell(pay_start, 1).value = "RECENT DEBT PAYMENTS"
    ws6.cell(pay_start, 1).font = _font(12, bold=True, color=M)
    pay_hdrs = ['Receipt', 'Customer', f'Amount ({currency})', 'Method', 'Date']
    style_header_row(ws6, pay_start + 1, pay_hdrs, [18, 24, 18, 14, 20])
    pay_rows = sorted(
        debt_payments, key=lambda x: x.get('created_at', ''), reverse=True
    )[:200]
    for i, pay in enumerate(pay_rows):
        r = pay_start + 2 + i
        vals = [
            pay.get('payment_receipt', ''),
            pay.get('customer_name', ''),
            float(pay.get('amount', 0) or 0),
            (pay.get('payment_method', '') or '').upper(),
            pay.get('created_at', '') or '',
        ]
        kinds = ['text', 'text', 'currency', 'center', 'datetime']
        for col, (val, kind) in enumerate(zip(vals, kinds), 1):
            apply_data_cell(
                ws6.cell(r, col), val, kind=kind, currency=currency, alt=i % 2 == 1)

    write_footer(
        ws6, pay_start + 4 + len(pay_rows), 9,
        record_count=len(inv_rows) + len(pay_rows),
        extra=f"Outstanding {outstanding:,.2f}",
    )
    if inv_rows:
        finalize_table(ws6, INV_HDR, INV_HDR + 1, INV_HDR + len(inv_rows), 6)

    # ── Sheet 7 – Payment Variance ────────────────────────────────────────────
    variance_rows = variance_rows or []
    variance_summary = variance_summary or {}
    ws7 = wb.create_sheet("Payment Variance")
    hdr7 = write_report_header(
        ws7, shop_name=shop_name, title="Payment Variance Report",
        ncols=12, period=period, generated_by=who, filters=filt, currency=currency,
    )
    _kpi_block(ws7, [
        (f'Extra Received ({currency})', float(variance_summary.get('extra_received') or 0), 'currency'),
        (f'Returned ({currency})', float(variance_summary.get('returned') or 0), 'currency'),
        (f'Deposits ({currency})', float(variance_summary.get('deposits') or 0) + float(variance_summary.get('advances') or 0), 'currency'),
        (f'Tips ({currency})', float(variance_summary.get('tips') or 0), 'currency'),
        (f'Transport ({currency})', float(variance_summary.get('transport') or 0), 'currency'),
    ], start_row=5)
    VR = 13
    vh = [
        'Date', 'Sale #', 'Cashier', 'Method', 'Sale Total', 'Received',
        'Excess', 'Handling', 'Returned', 'Deposit', 'Tip', 'Transport',
    ]
    style_header_row(ws7, VR, vh, [16, 16, 14, 10, 12, 12, 10, 14, 10, 10, 10, 12])
    for i, row in enumerate(variance_rows):
        r = VR + 1 + i
        vals = [
            row.get('created_at') or '',
            row.get('receipt_number', ''),
            row.get('cashier_name', ''),
            (row.get('payment_method') or '').upper(),
            float(row.get('sale_total') or 0),
            float(row.get('amount_received') or 0),
            float(row.get('excess_amount') or 0),
            (row.get('handling') or '').replace('_', ' ').title(),
            float(row.get('change_returned') or 0),
            float(row.get('deposit_amount') or 0) + float(row.get('advance_amount') or 0),
            float(row.get('tip_amount') or 0),
            float(row.get('transport_amount') or 0),
        ]
        kinds = [
            'datetime', 'text', 'text', 'center',
            'currency', 'currency', 'currency', 'text',
            'currency', 'currency', 'currency', 'currency',
        ]
        for col, (val, kind) in enumerate(zip(vals, kinds), 1):
            apply_data_cell(
                ws7.cell(r, col), val, kind=kind, currency=currency, alt=i % 2 == 1)
    write_footer(ws7, VR + 2 + len(variance_rows), 12, record_count=len(variance_rows))
    if variance_rows:
        finalize_table(ws7, VR, VR + 1, VR + len(variance_rows), 12)

    # ── Save ──────────────────────────────────────────────────────────────────
    if output_path is None:
        output_path = os.path.join(
            get_export_dir(),
            f"sales_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
        )
    os.makedirs(os.path.dirname(output_path) or '.', exist_ok=True)
    wb.save(output_path)
    return output_path


def export_sales_report_html(
    sales_data, items_by_sale, shop_name='My Shop',
    start_date=None, end_date=None, output_path=None,
    currency='KES', generated_by=None, filters=None,
):
    """
    Printable HTML sales report (browser Print → PDF). No PDF library required.
    Returns the written file path.
    """
    import html as _html

    period = f"{start_date or '—'} to {end_date or str(date.today())}"
    who = generated_by or 'System'
    filt = filters or f"Completed sales · {period}"
    total_rev = sum(float(s.get('total', 0) or 0) for s in sales_data)
    total_count = len(sales_data)

    def esc(v):
        return _html.escape('' if v is None else str(v))

    rows_html = []
    for idx, sale in enumerate(sales_data):
        sid = sale.get('id') or sale.get('sale_id')
        n_items = len(items_by_sale.get(sid, sale.get('items', [])))
        rows_html.append(
            '<tr>'
            f'<td>{idx + 1}</td>'
            f'<td>{esc(sale.get("receipt_number", ""))}</td>'
            f'<td>{esc(sale.get("created_at", "") or "")}</td>'
            f'<td>{esc(sale.get("cashier_name", ""))}</td>'
            f'<td class="num">{n_items or sale.get("item_count", "") or 0}</td>'
            f'<td class="num">{float(sale.get("total", 0) or 0):,.2f}</td>'
            f'<td>{esc((sale.get("payment_method", "") or "").upper())}</td>'
            '</tr>'
        )
    if not rows_html:
        rows_html.append(
            '<tr><td colspan="7" class="empty">No sales in this period.</td></tr>'
        )

    body = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="utf-8"/>
<title>{esc(shop_name)} — Sales Report</title>
<style>
  @page {{ margin: 12mm; }}
  body {{ font-family: Segoe UI, Arial, sans-serif; color: #1a1a1a; margin: 24px; }}
  h1 {{ font-size: 20px; margin: 0 0 4px; }}
  .meta {{ color: #555; font-size: 12px; margin-bottom: 16px; }}
  .kpis {{ display: flex; gap: 24px; margin: 12px 0 20px; }}
  .kpi {{ border: 1px solid #ddd; border-radius: 6px; padding: 10px 14px; min-width: 140px; }}
  .kpi .lbl {{ font-size: 11px; color: #666; text-transform: uppercase; }}
  .kpi .val {{ font-size: 18px; font-weight: 700; margin-top: 4px; }}
  table {{ width: 100%; border-collapse: collapse; font-size: 12px; }}
  th, td {{ border: 1px solid #ccc; padding: 6px 8px; text-align: left; }}
  th {{ background: #1e3a5f; color: #fff; }}
  tr:nth-child(even) td {{ background: #f7f9fc; }}
  .num {{ text-align: right; font-variant-numeric: tabular-nums; }}
  .empty {{ text-align: center; color: #888; }}
  .foot {{ margin-top: 18px; font-size: 11px; color: #666; }}
  @media print {{
    body {{ margin: 0; }}
    .kpi {{ break-inside: avoid; }}
  }}
</style>
</head>
<body>
  <h1>{esc(shop_name)} — Sales Report</h1>
  <div class="meta">
    Period: {esc(period)} · Generated by {esc(who)} ·
    {esc(datetime.now().strftime('%Y-%m-%d %H:%M'))}<br/>
    Filters: {esc(filt)} · Currency: {esc(currency)}
  </div>
  <div class="kpis">
    <div class="kpi"><div class="lbl">Transactions</div>
      <div class="val">{total_count}</div></div>
    <div class="kpi"><div class="lbl">Revenue ({esc(currency)})</div>
      <div class="val">{total_rev:,.2f}</div></div>
  </div>
  <table>
    <thead>
      <tr>
        <th>#</th><th>Receipt</th><th>Date</th><th>Cashier</th>
        <th>Items</th><th>Total ({esc(currency)})</th><th>Payment</th>
      </tr>
    </thead>
    <tbody>
      {''.join(rows_html)}
    </tbody>
  </table>
  <div class="foot">
    MBT POS printable report · Use browser Print → Save as PDF ·
    {app_version()}
  </div>
</body>
</html>
"""
    if output_path is None:
        output_path = os.path.join(
            get_export_dir(),
            f"sales_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.html",
        )
    os.makedirs(os.path.dirname(output_path) or '.', exist_ok=True)
    with open(output_path, 'w', encoding='utf-8') as f:
        f.write(body)
    return output_path


# Back-compat alias used by older callers
_footer = lambda ws, row, ncols: write_footer(ws, row, ncols, record_count=0)
