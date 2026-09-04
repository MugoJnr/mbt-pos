"""Focused tests for Super-Admin PIN gated spreadsheet exports.

No live shop DB mutation — isolated temp DB + tempfile workbooks only.
"""
from __future__ import annotations

import os
import sys
import tempfile
import unittest
from unittest.mock import MagicMock, patch

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
if ROOT not in sys.path:
    sys.path.insert(0, ROOT)


class ExportPinProtectionTests(unittest.TestCase):
    def setUp(self):
        from desktop.utils.export_security import clear_export_pin_session
        clear_export_pin_session()
        self._tmpdir = tempfile.TemporaryDirectory(ignore_cleanup_errors=True)
        self._db_path = os.path.join(self._tmpdir.name, 'export_pin.db')
        self._patches = [
            patch('mbt_paths.get_db_path', return_value=self._db_path),
            patch('desktop.utils.api_client.get_db_path', return_value=self._db_path),
        ]
        for p in self._patches:
            p.start()
        import desktop.utils.api_client as ac
        ac._SCHEMA_READY = False
        self.ac = ac
        self.api = ac.APIClient()
        self.api._role = 'admin'
        self.api._user_id = 1
        self.api._username = 'admin'

    def tearDown(self):
        from desktop.utils.export_security import clear_export_pin_session
        clear_export_pin_session()
        for p in self._patches:
            p.stop()
        self._tmpdir.cleanup()
        self.ac._SCHEMA_READY = False

    def _set_pin(self, pin: str):
        from desktop.utils.security import set_superadmin_pin
        self.assertTrue(set_superadmin_pin(pin, self.api))

    def test_no_pin_configured_blocks_export(self):
        from desktop.utils.export_security import (
            is_superadmin_pin_configured,
            require_superadmin_pin_for_export,
            PIN_NOT_CONFIGURED_MSG,
        )
        self.assertFalse(is_superadmin_pin_configured(self.api))
        with patch('PyQt5.QtWidgets.QMessageBox.warning') as warn:
            result = require_superadmin_pin_for_export(
                self.api, None, reason='test export')
        self.assertIsNone(result)
        warn.assert_called_once()
        args = warn.call_args[0]
        self.assertIn('PIN Required', args[1])
        self.assertIn('not configured', args[2].lower())
        self.assertIn('Security', args[2])
        self.assertIn('never', PIN_NOT_CONFIGURED_MSG.lower())

    def test_wrong_pin_blocks_export(self):
        from desktop.utils.export_security import require_superadmin_pin_for_export
        self._set_pin('654321')
        with patch(
            'desktop.utils.security.prompt_superadmin_pin',
            return_value='000000',
        ), patch('PyQt5.QtWidgets.QMessageBox.critical') as crit, patch(
            'desktop.utils.api_client._audit',
        ):
            result = require_superadmin_pin_for_export(
                self.api, None, reason='test export')
        self.assertIsNone(result)
        crit.assert_called()

    def test_correct_pin_writes_protected_workbook(self):
        from desktop.utils.export_security import require_superadmin_pin_for_export
        from backend.report_export_service import (
            new_workbook_sheet,
            save_workbook_password_protected,
            apply_workbook_password_protection,
        )
        from openpyxl import load_workbook

        pin = '111111'
        self._set_pin(pin)
        with patch(
            'desktop.utils.security.prompt_superadmin_pin',
            return_value=pin,
        ), patch('desktop.utils.api_client._audit'):
            got = require_superadmin_pin_for_export(
                self.api, None, reason='test export')
        self.assertEqual(got, pin)

        path = os.path.join(self._tmpdir.name, 'protected.xlsx')
        wb, ws = new_workbook_sheet('Sensitive')
        ws['A1'] = 'valuation'
        save_workbook_password_protected(wb, path, got)
        self.assertTrue(os.path.isfile(path))
        self.assertGreater(os.path.getsize(path), 0)

        loaded = load_workbook(path)
        self.assertTrue(bool(loaded.security and loaded.security.workbookPassword))
        self.assertTrue(loaded.security.lockStructure)
        self.assertTrue(loaded.active.protection.sheet)
        self.assertTrue(bool(loaded.active.protection.password))

        # Empty password must refuse (never unlock fallback)
        wb2, _ = new_workbook_sheet('Nope')
        with self.assertRaises(ValueError):
            apply_workbook_password_protection(wb2, '')

    def test_session_cache_skips_reprompt(self):
        from desktop.utils.export_security import (
            require_superadmin_pin_for_export,
            export_pin_session_active,
        )
        pin = '222222'
        self._set_pin(pin)
        with patch(
            'desktop.utils.security.prompt_superadmin_pin',
            return_value=pin,
        ) as prompt, patch('desktop.utils.api_client._audit'):
            first = require_superadmin_pin_for_export(self.api, None)
            second = require_superadmin_pin_for_export(self.api, None)
        self.assertEqual(first, pin)
        self.assertEqual(second, pin)
        self.assertEqual(prompt.call_count, 1)
        self.assertTrue(export_pin_session_active())

    def test_export_tabular_xlsx_with_password(self):
        from backend.report_export_service import export_tabular_xlsx
        from openpyxl import load_workbook

        path = os.path.join(self._tmpdir.name, 'tabular.xlsx')
        out = export_tabular_xlsx(
            title='Inventory Valuation',
            headers=['Item', 'Value'],
            rows=[['Widget', 100.0]],
            kinds=['text', 'currency'],
            shop_name='Test Shop',
            output_path=path,
            password='333333',
        )
        self.assertEqual(out, path)
        loaded = load_workbook(path)
        self.assertTrue(bool(loaded.security.workbookPassword))
        self.assertTrue(loaded.active.protection.sheet)

    def test_sales_export_engine_password_flag(self):
        try:
            from backend.export_engine import export_sales_report
        except ImportError as e:
            self.skipTest(f'export deps missing: {e}')
        from openpyxl import load_workbook
        from datetime import date

        path = os.path.join(self._tmpdir.name, 'sales.xlsx')
        export_sales_report(
            sales_data=[{
                'id': 1,
                'receipt_number': 'R-PIN-1',
                'created_at': f'{date.today()} 10:00:00',
                'cashier_name': 'admin',
                'subtotal': 50.0,
                'discount': 0,
                'tax': 0,
                'total': 50.0,
                'payment_method': 'Cash',
                'item_count': 1,
            }],
            items_by_sale={1: [{
                'product_name': 'Pin Item',
                'sku': 'P1',
                'quantity': 1,
                'unit_price': 50.0,
                'discount': 0,
                'total': 50.0,
            }]},
            shop_name='Pin Shop',
            start_date=str(date.today()),
            end_date=str(date.today()),
            output_path=path,
            password='444444',
        )
        loaded = load_workbook(path)
        self.assertTrue(bool(loaded.security.workbookPassword))
        self.assertTrue(any(ws.protection.sheet for ws in loaded.worksheets))


if __name__ == '__main__':
    unittest.main()
