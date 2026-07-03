"""
Import inventory from Excel into MBT POS SQLite database.

Usage:
  py tools/import_inventory.py "path/to/file.xlsx" --dry-run
  py tools/import_inventory.py "path/to/file.xlsx" --apply
  py tools/import_inventory.py "path/to/file.xlsx" --apply --db "C:\\...\\mbt_pos.db"
"""
from __future__ import annotations

import argparse
import os
import shutil
import sqlite3
import sys
from datetime import datetime

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
if ROOT not in sys.path:
    sys.path.insert(0, ROOT)

from mbt_paths import get_db_path


def _safe_int(v, default=0) -> int:
    if v is None or v == "":
        return default
    try:
        return int(float(v))
    except (TypeError, ValueError):
        return default


def _safe_float(v, default=0.0) -> float:
    if v is None or v == "":
        return default
    try:
        return float(v)
    except (TypeError, ValueError):
        return default


def _norm(s: str | None) -> str:
    return (s or "").strip().lower()


def load_rows(xlsx_path: str) -> list[dict]:
    try:
        import openpyxl
    except ImportError:
        raise SystemExit("Install openpyxl: py -m pip install openpyxl")

    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows: list[dict] = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue
        if not row or not row[2]:
            continue
        section = (str(row[0]).strip() if row[0] else "") or ""
        category = (str(row[1]).strip() if row[1] else "") or ""
        cat_full = category
        if section and category:
            cat_full = f"{section} — {category}"
        elif section:
            cat_full = section

        unit = row[3]
        unit_s = str(unit).strip() if unit not in (None, "") else "pcs"

        rows.append({
            "name": str(row[2]).strip(),
            "category": cat_full or "General",
            "unit": unit_s,
            "cost_price": _safe_float(row[4]),
            "stock": _safe_int(row[5]),
            "sku": str(row[6]).strip() if row[6] else None,
            "min_stock": max(0, _safe_int(row[7], 5)),
        })
    wb.close()
    return rows


def import_inventory(db_path: str, xlsx_path: str, apply: bool) -> dict:
    sheet_rows = load_rows(xlsx_path)
    if not sheet_rows:
        raise SystemExit("No product rows found in spreadsheet.")

    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    existing = conn.execute(
        "SELECT id, name, sku, stock, price, cost_price FROM products WHERE is_active=1"
    ).fetchall()
    by_sku = {_norm(r["sku"]): r for r in existing if r["sku"]}
    by_name = {_norm(r["name"]): r for r in existing}

    stats = {
        "sheet_rows": len(sheet_rows),
        "create": 0,
        "update": 0,
        "skip_dup_sku_in_sheet": 0,
        "errors": [],
    }

    seen_skus: set[str] = set()
    now = datetime.now().isoformat()

    for item in sheet_rows:
        sku_key = _norm(item["sku"])
        if sku_key:
            if sku_key in seen_skus:
                stats["skip_dup_sku_in_sheet"] += 1
                continue
            seen_skus.add(sku_key)

        match = None
        if sku_key and sku_key in by_sku:
            match = by_sku[sku_key]
        elif _norm(item["name"]) in by_name:
            match = by_name[_norm(item["name"])]

        try:
            if match:
                old_stock = int(match["stock"] or 0)
                new_stock = item["stock"]
                if apply:
                    conn.execute(
                        """UPDATE products SET name=?, sku=?, category=?, unit=?,
                           cost_price=?, stock=?, min_stock=?, updated_at=?
                           WHERE id=?""",
                        (
                            item["name"],
                            item["sku"],
                            item["category"],
                            item["unit"],
                            item["cost_price"],
                            new_stock,
                            item["min_stock"],
                            now,
                            match["id"],
                        ),
                    )
                    if new_stock != old_stock:
                        conn.execute(
                            """INSERT INTO stock_movements
                               (product_id, product_name, movement_type,
                                qty_before, qty_change, qty_after,
                                reference, reason, username)
                               VALUES (?,?,?,?,?,?,?,?,?)""",
                            (
                                match["id"],
                                item["name"],
                                "IMPORT_ADJUST",
                                old_stock,
                                new_stock - old_stock,
                                new_stock,
                                "SPREADSHEET_IMPORT",
                                "Agrovet Inventory.xlsx stock take",
                                "import_script",
                            ),
                        )
                stats["update"] += 1
            else:
                if apply:
                    conn.execute(
                        """INSERT INTO products
                           (name, sku, category, price, cost_price, stock, min_stock, unit)
                           VALUES (?,?,?,?,?,?,?,?)""",
                        (
                            item["name"],
                            item["sku"],
                            item["category"],
                            0.0,
                            item["cost_price"],
                            item["stock"],
                            item["min_stock"],
                            item["unit"],
                        ),
                    )
                    pid = conn.execute("SELECT last_insert_rowid()").fetchone()[0]
                    if item["stock"] > 0:
                        conn.execute(
                            """INSERT INTO stock_movements
                               (product_id, product_name, movement_type,
                                qty_before, qty_change, qty_after,
                                reference, reason, username)
                               VALUES (?,?,?,?,?,?,?,?,?)""",
                            (
                                pid,
                                item["name"],
                                "INITIAL",
                                0,
                                item["stock"],
                                item["stock"],
                                "SPREADSHEET_IMPORT",
                                "Agrovet Inventory.xlsx opening stock",
                                "import_script",
                            ),
                        )
                    new_row = conn.execute(
                        "SELECT id, name, sku, stock FROM products WHERE id=?",
                        (pid,),
                    ).fetchone()
                    if item["sku"]:
                        by_sku[_norm(item["sku"])] = new_row
                    # Prefer SKU matching when names repeat in the sheet.
                    if _norm(item["name"]) not in by_name or item["sku"]:
                        by_name[_norm(item["name"])] = new_row
                stats["create"] += 1
        except Exception as e:
            stats["errors"].append(f"{item.get('sku') or item['name']}: {e}")

    if apply:
        conn.commit()
    conn.close()
    return stats


def main():
    ap = argparse.ArgumentParser(description="Import MBT POS inventory from Excel")
    ap.add_argument("xlsx", help="Path to .xlsx file")
    ap.add_argument("--apply", action="store_true", help="Write to database (default: dry-run)")
    ap.add_argument("--dry-run", action="store_true", help="Preview only (default)")
    ap.add_argument("--db", help="Override database path")
    args = ap.parse_args()

    xlsx = os.path.abspath(args.xlsx)
    if not os.path.isfile(xlsx):
        raise SystemExit(f"File not found: {xlsx}")

    db_path = os.path.abspath(args.db) if args.db else get_db_path()
    apply = args.apply and not args.dry_run

    print(f"Spreadsheet: {xlsx}")
    print(f"Database:    {db_path}")
    print(f"Mode:        {'APPLY' if apply else 'DRY-RUN'}")
    print()

    if apply:
        backup = db_path + f".backup-{datetime.now().strftime('%Y%m%d-%H%M%S')}"
        shutil.copy2(db_path, backup)
        print(f"Backup:      {backup}")
        print()

    stats = import_inventory(db_path, xlsx, apply=apply)
    print("Results:")
    for k, v in stats.items():
        print(f"  {k}: {v}")
    if not apply:
        print()
        print("No changes written. Re-run with --apply to import.")


if __name__ == "__main__":
    main()
